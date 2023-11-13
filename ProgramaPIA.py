from datetime import date, datetime
import sys
import sqlite3
from sqlite3 import Error
import datetime
import openpyxl
import csv
import re

try:
    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
        mi_cursor=conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS clientes \
        (clave_cliente INTEGER PRIMARY KEY, nombre_cliente TEXT NOT NULL, RFC TEXT NOT NULL, correo_cliente TEXT NOT NULL, estatus TEXT NOT NULL);")

        mi_cursor.execute("CREATE TABLE IF NOT EXISTS servicios \
        (clave_servicio INTEGER PRIMARY KEY, nombre_servicio TEXT NOT NULL, costo_servicio NUMBER, estatus INTEGER NOT NULL);")

        mi_cursor.execute("CREATE TABLE IF NOT EXISTS notas (folio_nota INTEGER PRIMARY KEY, fecha_nota timestamp, clave_cliente INTEGER NOT NULL, total_nota NUMBER, estatus TEXT NOT NULL, FOREIGN KEY (clave_cliente) REFERENCES clientes (clave_cliente));")

        mi_cursor.execute("CREATE TABLE IF NOT EXISTS detalle_notas (id_detalle INTEGER PRIMARY KEY, folio_nota INTEGER NOT NULL, clave_servicio INTEGER NOT NULL, FOREIGN KEY (folio_nota) REFERENCES NOTAS (folio_nota), FOREIGN KEY (clave_servicio) REFERENCES servicios (clave_servicio));")

        print ('Tablas creadas exitosamente')

#############################CLIENTES
        def agregar_cliente():
            while True:
                try:
                    nombre_cliente=input("Ingrese el nombre del cliente: ").strip()
                    nombre_cliente=nombre_cliente.capitalize()
                    if not nombre_cliente.strip():
                        print('NO SE PUEDE QUEDAR EL NOMBRE DEL CLIENTE EN BLANCO, INTENTE DE NUEVO')
                        continue
                    if nombre_cliente.isdigit():
                        print('EL NOMBRE DEL CLIENTE NO PUEDE SER UN NÚMERO, INTENTE DE NUEVO')
                        continue
                    else:
                        break
                except ValueError:
                    print('Ingrese un nombre válido. Intentélo de nuevo.')
                    continue

            while True:
                try:
                    RFC=input("Ingrese el RFC del cliente (formato: ABCD123456XXX): ").strip().upper()
                    if RFC.strip() == '':
                        print("NO ES POSIBLE DEJAR EL RFC EN BLANCO, INTENTE DE NUEVO")
                        continue
                    if not re.match(r'^[A-Z&Ñ]{4}(0[0-9]|[1-9][0-9]|1[0-9][0-9]|2[0-3])(0[1-9]|1[0-2])(0[1-9]|[12][0-9]|3[01])[0-9A-Z]{3}$', RFC):
                        print(f"\nEL FORMATO DEL RFC ES INCORRECTO, INTENTE DE NUEVO (formato: ABCD123456XXX)")
                        continue
                    else:
                        break
                except ValueError:
                    print('Ingrese un nombre válido. Intentélo de nuevo.')
                    continue

            while True:
                try:
                    correo_cliente=input("Ingrese el correo electrónico del cliente (formato: tunombre@ejemplo.com): ")
                    validacion="(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"

                    if correo_cliente.strip() == '':
                        print("NO ES POSIBLE DEJAR EL CORREO EN BLANCO, INTENTE DE NUEVO")
                        continue

                    if not re.match(validacion, correo_cliente):
                        print("EL FORMATO DEL CORREO ES INCORRECTO, INTENTE DE NUEVO")
                        continue

                    else:
                        estatus= '1'
                        with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                            mi_cursor= conn.cursor()
                            valores=(nombre_cliente, RFC, correo_cliente, estatus)
                            mi_cursor.execute("INSERT INTO clientes (nombre_cliente, RFC, correo_cliente, estatus) VALUES (?,?,?,?)", valores)
                            conn.commit()

                            clave_cliente=mi_cursor.lastrowid
                            print(f'Se registró el cliente con la clave: {clave_cliente}')
                            break

                except ValueError:
                    print('Ingrese un correo válido. Intentélo de nuevo.')
                    continue
        def suspender_cliente():
          try:
            with sqlite3.connect('Evidencia3_Prueba.db') as conn:
              mi_cursor = conn.cursor()

              mi_cursor.execute("SELECT clave_cliente, nombre_cliente FROM clientes WHERE estatus='1'")
              clientes_activos = mi_cursor.fetchall()

              if not clientes_activos:
                print("No hay clientes activos para suspender.")
                return

              print("\nReporte de Clientes Activos:")
              print("Clave cliente\t\tNombre cliente")
              for cliente in clientes_activos:
                print(f"\t{cliente[0]}\t\t{cliente[1]}")

              clave_suspender = input("\nIngrese la clave del cliente que desea suspender (0 para volver al menú anterior): ").strip()
              if clave_suspender == '0':
                print("Volviendo al menú anterior.")
                return

              mi_cursor.execute("SELECT * FROM clientes WHERE clave_cliente=? AND estatus='1'", (clave_suspender,))
              cliente_suspender = mi_cursor.fetchone()

              if not cliente_suspender:
                print("Clave de cliente no válida.")
                return

              # Mostrar los datos del cliente y solicitar confirmación
              print("\nDatos del cliente a suspender:")
              print(f"Clave cliente: {cliente_suspender[0]}")
              print(f"Nombre cliente: {cliente_suspender[1]}")
              print(f"RFC: {cliente_suspender[2]}")
              print(f"Correo cliente: {cliente_suspender[3]}")

              confirmacion = input("\n¿Desea suspender a este cliente? (Sí/No): ").strip().lower()
              if confirmacion == 'si':
                # Actualizar el estatus del cliente a suspendido
                mi_cursor.execute("UPDATE clientes SET estatus='0' WHERE clave_cliente=?", (clave_suspender,))
                conn.commit()
                print(f"El cliente con clave {cliente_suspender[0]} ha sido suspendido.")
              else:
                print("Operación cancelada. Volviendo al menú anterior.")

          except Exception as e:
            print(f"Ocurrió un error: {str(e)}")


        def recuperar_cliente():
          try:
           with sqlite3.connect('Evidencia3_Prueba.db') as conn:
            mi_cursor = conn.cursor()

            mi_cursor.execute("SELECT clave_cliente, nombre_cliente FROM clientes WHERE estatus = '0'")
            clientes_suspendidos = mi_cursor.fetchall()

            if not clientes_suspendidos:
              print("No hay clientes suspendidos para recuperar")
              return

            print("\nReporte de Clientes Suspendidos:")
            print("Clave cliente\t\tNombre cliente")
            for cliente in clientes_suspendidos:
             print(f"\t{cliente[0]}\t\t{cliente[1]}")

        # Solicitar al usuario que elija una clave de cliente para recuperar
            clave_recuperar = input("\nIngrese la clave del cliente que desea recuperar (0 para volver al menú anterior): ").strip()

            if clave_recuperar == '0':
             print("Volviendo al menú anterior.")
             return
            mi_cursor.execute("SELECT * FROM clientes WHERE clave_cliente=? AND estatus='0'", (clave_recuperar,))
            cliente_recuperar = mi_cursor.fetchone()

            if not cliente_recuperar:
             print("Clave de cliente no válida.")
             return
        # Mostrar los datos del cliente y solicitar confirmación
            print("\nDatos del cliente a recuperar:")
            print(f"Clave cliente: {cliente_recuperar[0]}")
            print(f"Nombre cliente: {cliente_recuperar[1]}")
            print(f"RFC: {cliente_recuperar[2]}")
            print(f"Correo cliente: {cliente_recuperar[3]}")
            confirmacion = input("\n¿Desea recuperar a este cliente? (Sí/No): ").strip().lower()

            if confirmacion == 'si' or confirmacion == 'SI':
            # Actualizar el estatus del cliente a activo
             mi_cursor.execute("UPDATE clientes SET estatus='1' WHERE clave_cliente=?", (clave_recuperar,))
             conn.commit()
             print(f"El cliente con clave {cliente_recuperar[0]} ha sido recuperado.")
            else:
              print("Operación cancelada. Volviendo al menú anterior.")
          except Exception as e:
           print(f"Ocurrió un error: {str(e)}")
        #consultas y reportes de clientes
        #LISTADO DE clientes
        def ordenar_cliente_por_clave():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        print(f"\nClientes ordenados por clave: ")
                        mi_cursor.execute(f"SELECT * FROM clientes WHERE estatus='1' ORDER BY clave_cliente")
                        clientes_orden_clave=mi_cursor.fetchall()
                        print(f"\nClave cliente\t\tNombre cliente\t\tRFC\t\t\tCorreo cliente")
                        for cliente in clientes_orden_clave:
                            print(f"\t{cliente[0]}\t\t{cliente[1]}\t\t{cliente[2]}\t\t{cliente[3]}")

                        print(f"\nOpciones de exportación:")
                        print(f"1. Excel")
                        print(f"2. CSV")
                        print(f"3. Volver a menú de reportes")

                        opcion_exportar=int(input(f'\nOpción a elegir: '))
                        match opcion_exportar:
                            case 1:
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                sheet.title = "Clientes_Ordenados_Clave"

                                sheet['A1'] = "Clave cliente"
                                sheet['B1'] = "Nombre cliente"
                                sheet['C1'] = "RFC"
                                sheet['D1'] = "Correo cliente"

                                row = 2
                                for cliente in clientes_orden_clave:
                                    sheet[f'A{row}'] = cliente[0]
                                    sheet[f'B{row}'] = cliente[1]
                                    sheet[f'C{row}'] = cliente[2]
                                    sheet[f'd{row}'] = cliente[3]
                                    row+=1

                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_excel = f"ReporteClientesActivosPorClave_{fecha_reporte}.xlsx"
                                workbook.save(nombre_archivo_excel)
                                print(f"Se ha exportado la información a '{nombre_archivo_excel}'.")
                                break

                            case 2:
                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_csv=f"ReporteClientesActivosPorClave_{fecha_reporte}.csv"
                                with open(nombre_archivo_csv, 'w', newline='') as file:
                                    writer = csv.writer(file)
                                    writer.writerow(["Clave cliente", "Nombre cliente", "RFC", "Correo cliente"])
                                    for cliente in clientes_orden_clave:
                                        writer.writerow([cliente[0], cliente[1], cliente[2], cliente[3]])

                                    print(f"Se ha exportado la información a '{nombre_archivo_csv}'.")
                                    break

                            case 3:
                                print('VOLVIENDO AL MENÚ DE REPORTES.')
                                break
                            case _:
                                print("OPCIÓN NO VÁLIDA. INGRESE UN NÚMERO DEL 1 AL 3.")

                except ValueError:
                    print('INGRESE UNA OPCIÓN VÁLIDA')

                except Exception as e:
                    print(f"Ocurrió un error: {str(e)}")


        def ordenar_cliente_por_nombre():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        print(f"\nClientes ordenados por nombre: ")
                        mi_cursor.execute(f"SELECT * FROM clientes WHERE estatus='1' ORDER BY nombre_cliente")
                        clientes_orden_clave=mi_cursor.fetchall()
                        print(f"\nClave cliente\t\tNombre cliente\t\tRFC\t\t\tCorreo cliente")
                        for cliente in clientes_orden_clave:
                            print(f"\t{cliente[0]}\t\t{cliente[1]}\t\t{cliente[2]}\t\t{cliente[3]}")

                        print(f"\nOpciones de exportación:")
                        print(f"1. Excel")
                        print(f"2. CSV")
                        print(f"3. Volver a menú de reportes")

                        opcion_exportar=int(input(f'\nOpción a elegir: '))
                        match opcion_exportar:
                            case 1:
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                sheet.title = "Clientes_Ordenados_Clave"

                                sheet['A1'] = "Clave cliente"
                                sheet['B1'] = "Nombre cliente"
                                sheet['C1'] = "RFC"
                                sheet['D1'] = "Correo cliente"

                                row = 2
                                for cliente in clientes_orden_clave:
                                    sheet[f'A{row}'] = cliente[0]
                                    sheet[f'B{row}'] = cliente[1]
                                    sheet[f'C{row}'] = cliente[2]
                                    sheet[f'd{row}'] = cliente[3]
                                    row+=1

                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_excel = f"ReporteClientesActivosPorNombre_{fecha_reporte}.xlsx"
                                workbook.save(nombre_archivo_excel)
                                print(f"Se ha exportado la información a '{nombre_archivo_excel}'.")
                                break

                            case 2:
                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_csv=f"ReporteClientesActivosPorNombre_{fecha_reporte}.csv"
                                with open(nombre_archivo_csv, 'w', newline='') as file:
                                    writer = csv.writer(file)
                                    writer.writerow(["Clave cliente", "Nombre cliente", "RFC", "Correo cliente"])
                                    for cliente in clientes_orden_clave:
                                        writer.writerow([cliente[0], cliente[1], cliente[2], cliente[3]])

                                    print(f"Se ha exportado la información a '{nombre_archivo_csv}'.")
                                    break

                            case 3:
                                print('VOLVIENDO AL MENÚ DE REPORTES.')
                                break
                            case _:
                                print("OPCIÓN NO VÁLIDA. INGRESE UN NÚMERO DEL 1 AL 3.")

                except ValueError:
                    print('INGRESE UNA OPCIÓN VÁLIDA')

                except Exception as e:
                    print(f"Ocurrió un error: {str(e)}")

        def buscar_cliente_por_clave(clave):
            with sqlite3.connect('Evidencia3_Prueba.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT * FROM clientes WHERE clave_cliente = ?", (clave,))
                cliente = mi_cursor.fetchone()

                if cliente:
                    print(f"Cliente encontrado por clave {clave}:")
                    print("{:<15} {:<20} {:<15} {:<20}".format("Clave cliente", "Nombre cliente", "RFC", "Correo cliente"))
                    print("{:<15} {:<20} {:<15} {:<20}".format(cliente[0], cliente[1], cliente[2], cliente[3]))
                else:
                    print("No se encontró ningún cliente con esa clave.")

        def buscar_cliente_por_nombre(nombre):
            with sqlite3.connect('Evidencia3_Prueba.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT * FROM clientes WHERE nombre_cliente = ?", (nombre,))
                clientes = mi_cursor.fetchall()

                if clientes:
                    print(f"Clientes encontrados por nombre '{nombre}':")
                    print("{:<15} {:<20} {:<15} {:<20}".format("Clave cliente", "Nombre cliente", "RFC", "Correo cliente"))
                    for cliente in clientes:
                        print("{:<15} {:<20} {:<15} {:<20}".format(cliente[0], cliente[1], cliente[2], cliente[3]))
                else:
                    print("No se encontraron clientes con ese nombre.")

        def agregar_servicio():
            while True:
                try:
                    nombre_servicio=input("Ingrese el nombre del servicio: ").strip()
                    nombre_servicio=nombre_servicio.capitalize()
                    if not nombre_servicio.strip():
                        print('NO SE PUEDE QUEDAR EL NOMBRE DEL SERVICIO EN BLANCO, INTENTE DE NUEVO')
                        continue
                    if nombre_servicio.isdigit():
                        print('EL NOMBRE DEL SERVICIO NO PUEDE SER UN NÚMERO, INTENTE DE NUEVO')
                        continue
                    else:
                        break
                except ValueError:
                    print('Ingrese un nombre válido. Intentélo de nuevo.')
                    continue


            while True:
                try:
                    costo_servicio=int(input('Ingrese el costo del servicio: '))
                    if costo_servicio == 0:
                        print('EL COSTO DEL SERVICIO NO PUEDE SER 0. INTENTE DE NUEVO')
                        continue
                    if not costo_servicio:
                        print('EL COSTO DEL SERVICIO NO PUEDE QUEDAR EN BLANCO. INTENTE DE NUEVO')
                        continue
                    elif costo_servicio <= 0:
                        print('EL COSTO DEL SERVICIO DEBE SER MAYOR A 0. INTENTE DE NUEVO')
                        continue
                    else:
                        with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                            mi_cursor= conn.cursor()
                            valores=(nombre_servicio, costo_servicio, 1)
                            mi_cursor.execute("INSERT INTO servicios (nombre_servicio, costo_servicio, estatus) VALUES (?,?,?)", valores)
                            conn.commit()

                            clave_servicio=mi_cursor.lastrowid
                            print(f'Se registró el servicio con la clave: {clave_servicio}')
                            break

                except ValueError:
                    print('Ingrese un costo válido. Intentélo de nuevo.')
                    continue


        def servicios_suspender():
          try:
            with sqlite3.connect('Evidencia3_Prueba.db') as conn:
                 mi_cursor = conn.cursor()
                 print("\nServicios:")
                 mi_cursor.execute("SELECT clave_servicio, nombre_servicio FROM servicios")
                 servicios_activos = mi_cursor.fetchall()

                 if not servicios_activos:
                     print("No se encontraron servicios para suspender por el momento.")
                     return

                 print("\nClave servicio\t\tNombre servicio")
                 for servicio in servicios_activos:
                     print(f"\t{servicio[0]}\t\t{servicio[1]}")

            while True:
                try:
                    clave_servicio_suspender = int(input("Ingrese la clave del servicio que quieres suspender o 0 para regresar al menú anterior."))
                    if clave_servicio_suspender == 0:
                        print("Regresando al menú anterior.")
                        return
                    elif mi_cursor.execute("SELECT COUNT (*) FROM servicios WHERE clave_servicio =? AND estatus= '1'", (clave_servicio_suspender,)).fetchone()[0] == 0:
                        print("La clave es incorrecta o no se encuentra disponible. Intente de nuevo")
                    else:
                        break
                except ValueError:
                    print("Ingrese una opción válida.")
                    mi_cursor.execute("SELECT * FROM servicios WHERE clave_servicio=?", (clave_servicio_suspender,))
                    servicio_suspender = mi_cursor.fetchone()
                    print("\tDatos del servicio a suspender:")
                    print(f"\nClave servicio\t\tNombre servicio\t\tCosto servicio")
                    print(f"\t{servicio_suspender[0]}\t\t{servicio_suspender[1]}\t\t{servicio_suspender[2]}")

            while True:
               confirmacion = input("\n¿Desea suspender este servicio? (Si/No):").strip().lower()
               if confirmacion == 'si' or confirmacion == 's':
                  mi_cursor.execute("UPDATE servicios SET estatus= '0' WHERE clave_servicio=?", (clave_servicio_suspender,))
                  conn.commit()
                  print(f"El servicio con clave {clave_servicio_suspender} se ha suspendido.")
                  return
               elif confirmacion == 'no' or confirmacion == 'n':
                    print("Se canceló la acción. Volviendo al menú anterior.")
                    return
               else:
                    print("Ingrese 'Si' o 'No'.")
          except Exception as e:
               print(f"Ocurrió un error: {str(e)}")


        def recuperar_servicio():
          try:
            with sqlite3.connect('Evidencia3_Prueba.db') as conn:
               mi_cursor = conn.cursor()
               print("\nServicios suspendidos")
               mi_cursor.execute("SELECT clave_servicio, nombre_servicio, costo_servicio FROM servicios WHERE estatus= '0'")
               servicios_suspendidos = mi_cursor.fetchall()

               if not servicios_suspendidos:
                  print("No se encuentran servicios suspendidos por el momento.")
                  return

               print("\nClave servicio\t\tNombre servicio\t\tCosto servicio")
               for servicio in servicios_suspendidos:
                   print(f"\t{servicio[0]}\t\t{servicio[1]}\t\t{servicio[2]}")

               while True:
                try:
                    clave_servicio_recuperar = int(input("Ingrese la clave del servicio a recuperar o 0 para regresar al menú anterior."))
                    if clave_servicio_recuperar == 0:
                        print("Volviendo al menú anterior.")
                        return
                    elif mi_cursor.execute("SELECT COUNT(*) FROM servicios WHERE clave_servicio=? AND estatus='0'", (clave_servicio_recuperar,)).fetchone()[0] == 0:
                        print("La clave es incorrecta o el servicio no se encuentra suspendido. Intente de nuevo.")
                    else:
                        break
                except ValueError:
                    print("Ingrese una opción válida.")

               mi_cursor.execute("SELECT * FROM servicios WHERE clave_servicio =?", (clave_servicio_recuperar,))
               servicio_recuperar = mi_cursor.fetchone()

               print("\nDatos del servicio a recuperar.")
               print("\nClave servicio\t\tNombre servicio\t\tCosto servicio")
               print(f"\t{servicio_recuperar[0]}\t\t{servicio_recuperar[1]}\t\t{servicio_recuperar[2]}")

               while True:
                  confirmacion = input("\n¿Desea recuperar el servicio suspendido? (Si/No):").strip().lower()
                  if confirmacion == 'si':
                     mi_cursor.execute("UPDATE servicios SET estatus= '1' WHERE clave_servicio=?", (clave_servicio_recuperar,))
                     conn.commit()
                     print(f"\nEl servicio con clave {clave_servicio_recuperar} se ha recuperado.")
                     return
                  elif confirmacion == 'no':
                       print("Se canceló la acción. Volviendo al menú anterior.")
                       return
                  else:
                       print("Ingrese 'Si' o 'No'.")
          except Exception as e:
                 print(f"Ocurrió un error: {str(e)}")

        
        def buscar_servicio_por_clave():
            print(f'\nLista de servicios: ')
            with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                mi_cursor= conn.cursor()
                mi_cursor.execute("SELECT clave_servicio, nombre_servicio FROM servicios WHERE estatus= '1'")
                lista_servicios=mi_cursor.fetchall()
                print(f"\nClave servicio\tNombre servicio")
                for servicio in lista_servicios:
                    print(f"\t{servicio[0]}\t{servicio[1]}")
            while True:
                try:
                    clave_servicio=int(input(f'\nIngrese la clave del servicio a consultar / Si no se va consultar un servicio escriba 0: '))
                    if clave_servicio==0:
                        print('USTED HA DECIDIDO NO CONSULTAR NINGUN SERVICIO')
                        break
                    if not clave_servicio:
                        print('LA CLAVE NO SE PUEDE QUEDAR EN BLANCO. INGRESE UNA CLAVE VÁLIDA')
                        continue
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        mi_cursor.execute(f"SELECT * FROM servicios WHERE clave_servicio={clave_servicio}")
                        servicios_clave=mi_cursor.fetchall()
                        if clave_servicio not in [servicio[0] for servicio in servicios_clave]:
                            print('El servicio no se encuentra en el sistema. Por favor, ingrese una clave válida.')
                            continue
                        else:
                            print(f"\nClave servicio\t\tNombre servicio\t\tCosto servicio")
                            for servicio in servicios_clave:
                                print(f"\t{servicio[0]}\t\t{servicio[1]}\t\t{servicio[2]}")
                            break

                except ValueError:
                    print('Ingrese una clave válida. Intente de nuevo')

        def buscar_servicio_por_nombre():
            while True:
                try:
                    nombre_servicio_buscar=input(f'\nIngrese el nombre del servicio a buscar: ').strip()
                    nombre_servicio=nombre_servicio_buscar.capitalize()

                    if not nombre_servicio.strip():
                        print('NO SE PUEDE QUEDAR EL NOMBRE DEL SERVICIO EN BLANCO, INTENTE DE NUEVO')
                        continue
                    if nombre_servicio.isdigit():
                        print('EL NOMBRE DEL SERVICIO NO PUEDE SER UN NÚMERO, INTENTE DE NUEVO')
                        continue
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        mi_cursor.execute(f"SELECT * FROM servicios WHERE nombre_servicio='{nombre_servicio} AND estatus= '1''")
                        servicios_nombre=mi_cursor.fetchall()
                        if nombre_servicio not in [servicio[1] for servicio in servicios_nombre]:
                            print('EL SERVICIO NO SE ENCUENTRA EN EL SISTEMA. INGRESE UN NOMBRE VÁLIDO')
                            continue
                        else:
                            print(f"\nClave servicio\t\tNombre servicio\t\tCosto servicio")
                            for servicio in servicios_nombre:
                                print(f"\t{servicio[0]}\t\t{servicio[1]}\t\t${servicio[2]}")
                            break
                except ValueError:
                    print('INGRESE UN NOMBRE VÁLIDO')

            
        def ordenar_servicio_por_clave():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        print(f"\nServicios ordenados por clave: ")
                        mi_cursor.execute(f"SELECT * FROM servicios WHERE estatus= '1' ORDER BY clave_servicio")
                        servicios_orden_clave=mi_cursor.fetchall()
                        print(f"\nClave servicio\t\tNombre servicio\t\tCosto servicio")
                        for servicio in servicios_orden_clave:
                            print(f"\t{servicio[0]}\t\t{servicio[1]}\t\t${servicio[2]}")

                        print(f"\nOpciones de exportación:")
                        print(f"1. Excel")
                        print(f"2. CSV")
                        print(f"3. Volver a menú de reportes")

                        opcion_exportar=int(input(f'\nOpción a elegir: '))
                        match opcion_exportar:
                            case 1:
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                sheet.title = "Servicios_Ordenados_Clave"

                                sheet['A1'] = "Clave servicio"
                                sheet['B1'] = "Nombre servicio"
                                sheet['C1'] = "Costo servicio"

                                row = 2
                                for servicio in servicios_orden_clave:
                                    sheet[f'A{row}'] = servicio[0]
                                    sheet[f'B{row}'] = servicio[1]
                                    sheet[f'C{row}'] = servicio[2]
                                    row+=1

                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_excel = f"ReporteServiciosPorClave_{fecha_reporte}.xlsx"
                                workbook.save(nombre_archivo_excel)
                                print(f"Se ha exportado la información a '{nombre_archivo_excel}'.")
                                break

                            case 2:
                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_csv=f"ReporteServiciosPorClave_{fecha_reporte}.csv"
                                with open(nombre_archivo_csv, 'w', newline='') as file:
                                    writer = csv.writer(file)
                                    writer.writerow(["Clave servicio", "Nombre servicio", "Costo servicio"])
                                    for servicio in servicios_orden_clave:
                                        writer.writerow([servicio[0], servicio[1], servicio[2]])

                                    print(f"Se ha exportado la información a '{nombre_archivo_csv}'.")
                                    break
                            case 3:
                                print('VOLVIENDO AL MENÚ DE REPORTES.')
                                break
                            case _:
                                print("OPCIÓN NO VÁLIDA. INGRESE UN NÚMERO DEL 1 AL 3.")

                except ValueError:
                    print('INGRESE UNA OPCIÓN VÁLIDA')

                except Exception as e:
                    print(f"Ocurrió un error: {str(e)}")


        def ordenar_servicio_por_nombre():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        print(f"\nServicios ordenados por nombre: ")
                        mi_cursor.execute(f"SELECT * FROM servicios WHERE estatus= '1' ORDER BY nombre_servicio")
                        servicios_orden_nombre=mi_cursor.fetchall()
                        print(f"\nClave servicio\t\tNombre servicio\t\tCosto servicio")
                        for servicio in servicios_orden_nombre:
                            print(f"\t{servicio[0]}\t\t{servicio[1]}\t\t${servicio[2]}")

                        print(f"\nOpciones de exportación:")
                        print(f"1. Excel")
                        print(f"2. CSV")
                        print(f"3. Volver a menú de reportes")

                        opcion_exportar=int(input(f'\nOpción a elegir: '))
                        match opcion_exportar:
                            case 1:
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                sheet.title = "Servicios_Ordenados_Clave"

                                sheet['A1'] = "Clave servicio"
                                sheet['B1'] = "Nombre servicio"
                                sheet['C1'] = "Costo servicio"

                                row = 2
                                for servicio in servicios_orden_nombre:
                                    sheet[f'A{row}'] = servicio[0]
                                    sheet[f'B{row}'] = servicio[1]
                                    sheet[f'C{row}'] = servicio[2]
                                    row+=1

                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_excel = f"ReporteServiciosPorNombre_{fecha_reporte}.xlsx"
                                workbook.save(nombre_archivo_excel)
                                print(f"Se ha exportado la información a '{nombre_archivo_excel}'.")
                                break

                            case 2:
                                fecha_actual=datetime.date.today()
                                fecha_reporte = fecha_actual.strftime("%m-%d-%Y")

                                nombre_archivo_csv=f"ReporteServiciosPorNombre_{fecha_reporte}.csv"
                                with open(nombre_archivo_csv, 'w', newline='') as file:
                                    writer = csv.writer(file)
                                    writer.writerow(["Clave servicio", "Nombre servicio", "Costo servicio"])
                                    for servicio in servicios_orden_nombre:
                                        writer.writerow([servicio[0], servicio[1], servicio[2]])

                                    print(f"Se ha exportado la información a '{nombre_archivo_csv}'.")
                                    break
                            case 3:
                                print('VOLVIENDO AL MENÚ DE REPORTES.')
                                break
                            case _:
                                print("OPCIÓN NO VÁLIDA. INGRESE UN NÚMERO DEL 1 AL 3.")

                except ValueError:
                    print('INGRESE UNA OPCIÓN VÁLIDA')

                except Exception as e:
                    print(f"Ocurrió un error: {str(e)}")


 ############################ NOTAS
        def capturar_fecha():
            while True:
                fecha_ingresada= input('Fecha de la nota a capturar (dd/mm/YYYY): ')
                try:
                    fecha=datetime.datetime.strptime(fecha_ingresada, "%d/%m/%Y").date()
                except ValueError:
                    print('FORMATO DE FECHA INCORRECTO. INTENTE DE NUEVO.')
                    continue

                hoy_fecha = datetime.date.today()
                if (fecha.month, fecha.day) >= (hoy_fecha.month, hoy_fecha.day):
                    print('LA FECHA INGRESADA NO PUEDE SER MAYOR AL DÍA DE HOY. INTENTE DE NUEVO')
                    continue
                else:
                    return fecha



        def registrar_nota():
            fecha_nota=capturar_fecha()

            print(f'\nLista de clientes: ')
            with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                conn.execute("PRAGMA foreign_keys=1")
                mi_cursor= conn.cursor()
                mi_cursor.execute("SELECT clave_cliente, nombre_cliente FROM clientes")
                listaclientes=mi_cursor.fetchall()
                for clave, nombre in listaclientes:
                    print (f"Clave cliente: {clave}\tNombre cliente: {nombre}")
            while True:
                try:
                    clave_cliente=int(input('Ingrese la clave del cliente: '))
                    if clave_cliente == 0:
                        print('LA CLAVE DEL CLIENTE NO PUEDE SER 0. INTENTE DE NUEVO.')
                        continue
                    if not clave_cliente:
                        print('LA CLAVE DEL CLIENTE NO PUEDE QUEDAR EN BLANCO. INGRESE UNA CLAVE VÁLIDA')
                        continue
                    elif clave_cliente not in [clave[0] for clave in listaclientes]:
                        print('LA CLAVE SELECCIONADA NO ESTÁ REGISTRADA. INTENTE DE NUEVO')
                        continue

                except ValueError:
                    print('INGRESE UN NÚMERO VÁLIDO. INTENTE DE NUEVO')
                    continue
                break

            servicios_en_nota=[]
            total_nota=0
            while True:
                print (f'\nLista de servicios: ')
                with sqlite3.connect ('Evidencia3_Prueba.db') as conn:

                    mi_cursor= conn.cursor()
                    mi_cursor.execute("SELECT clave_servicio, nombre_servicio, costo_servicio FROM servicios WHERE estatus= '1'")
                    listaservicios=mi_cursor.fetchall()

                    for clave, nombre, costo in listaservicios:
                        print (f"Clave servicio: {clave}\tNombre servicio: {nombre}\tCosto servicio: {costo}")

                    try:
                        clave_servicio=int(input('Ingrese la clave del servicio a agregar o escriba 0 si ya no desea agregar servicios: '))
                        if clave_servicio == 0:
                            print('HA DECIDIO NO AGREGAR MÁS SERVICIOS')
                            break

                        elif clave_servicio not in [servicio[0] for servicio in listaservicios]:
                            print('LA CLAVE SELECCIONADA NO ESTÁ REGISTRADA. INTENTE DE NUEVO')
                            continue

                    except ValueError:
                        print('INGRESE UN NÚMERO VÁLIDO. INTENTE DE NUEVO.')
                        continue

                    servicios_en_nota.append(clave_servicio)

                    mi_cursor.execute(f"SELECT costo_servicio FROM servicios WHERE clave_servicio={clave_servicio}")
                    listacostos=mi_cursor.fetchall()
                    for costo in listacostos:
                        total_nota= total_nota + int(costo[0])

                    estatus='1'

            with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                mi_cursor= conn.cursor()

                print(f'\nTotal servicios: {total_nota}')
                valores=(fecha_nota, clave_cliente, total_nota, estatus)
                mi_cursor.execute("INSERT INTO notas (fecha_nota, clave_cliente, total_nota, estatus) VALUES (?,?,?,?)", valores)
                conn.commit()

                folio_nota=mi_cursor.lastrowid

                for clave_servicio in servicios_en_nota:
                    valores=(folio_nota, clave_servicio)
                    mi_cursor.execute("INSERT INTO detalle_notas (folio_nota, clave_servicio) VALUES (?,?)", valores)

                print(f"\nNota registrada correctamente con el folio: {folio_nota}")

        def cancelar_nota():
            while True:
                try:
                    folio_cancelado = int(input('Ingrese el folio de la nota a cancelar: '))
                    if not folio_cancelado:
                        print('LA OPCIÓN ESTÁ EN BLANCO. INTENTE DE NUEVO.')
                        continue
                except ValueError:
                    print('INGRESE UN NÚMERO VÁLIDO. INTENTE DE NUEVO')
                    continue

                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        mi_cursor.execute(f"SELECT * FROM notas WHERE folio_nota={folio_cancelado} AND estatus='1' ")
                        folios_cancelados=mi_cursor.fetchall()
                        if not folios_cancelados:
                            print('EL FOLIO DE LA NOTA A CANCELAR NO EXISTE. INTENTE DE NUEVO.')
                            continue
                        else:
                            print(f"\nFolio nota\tFecha nota\tClave cliente\tTotal nota\tEstatus")
                            for folio in folios_cancelados:
                                print(f"{folio[0]}\t\t{folio[1]}\t\t{folio[2]}\t{folio[3]}\t{folio[4]}")
                            try:
                                opcion_cancelar=input('Desea cancelar la nota? SI/NO: ').upper().strip()
                                if not opcion_cancelar:
                                    print('LA OPCIÓN ESTÁ EN BLANCO. INTENTE DE NUEVO')
                                    continue
                                if opcion_cancelar=="SI":
                                    mi_cursor.execute(f"UPDATE notas SET estatus='0' WHERE folio_nota={folio_cancelado}")
                                    conn.commit()
                                    print('SE CANCELÓ LA NOTA EXITOSAMENTE')
                                    break
                                elif opcion_cancelar == "NO":
                                    print('NO SE CANCELÓ LA NOTA SELECCIONADA')
                                    break
                                else:
                                    print('Opción inválida. Escriba "SI" o "NO". Inténtelo de nuevo.')
                                    continue
                            except ValueError:
                                print('Opción inválida. Escriba "SI" o "NO". Inténtelo de nuevo.')

                except Exception as e:
                    print('Ocurrió un error: {e} ')
                    continue
                except sqlite3.Error as e:
                    print(f'Ocurrió un error de base de datos: {e}')
                    break

        def recuperar_nota():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()
                        mi_cursor.execute(f"SELECT folio_nota FROM notas WHERE estatus='0' ")
                        folios_cancelados=mi_cursor.fetchall()
                        if not folios_cancelados:
                            print('NO HAY NOTAS CANCELADAS PARA MOSTRAR')
                            break
                        else:
                            print(f"\nFolios de las notas actualmente canceladas:")
                            for folio in folios_cancelados:
                                print(f"\t{folio[0]}")

                            try:
                                folio_recuperar=int(input('Ingrese el folio de la nota a recuperar / Si no se va a recuperar una nota escriba 0: '))
                                if folio_recuperar == 0:
                                    print('EL FOLIO DE LA NOTA NO PUEDE SER 0. INGRESE UN FOLIO VÁLIDO')
                                    continue
                                if not folio_recuperar:
                                    print('EL FOLIO NO SE PUEDE QUEDAR EN BLANCO. INGRESE UN FOLIO VÁLIDO')
                                    continue
                                if folio_recuperar not in [folio[0] for folio in folios_cancelados]:
                                    print('EL FOLIO INGRESADO NO EXISTE EN LAS NOTAS CANCELADAS. INGRESE UN FOLIO VÁLIDO')
                                    continue
                            except ValueError:
                                print('Ingrese un número válido. Inténtelo de nuevo.')
                            else:
                                opcion_recuperar=input('Desea recuperar la nota? SI/NO: ').upper()
                                if not opcion_recuperar:
                                    print('LA OPCIÓN ESTÁ EN BLANCO. INTENTE DE NUEVO')
                                    continue
                                if opcion_recuperar=='SI':
                                    for folio in folios_cancelados:
                                        if folio_recuperar==folio[0]:
                                            mi_cursor.execute(f"UPDATE notas SET estatus='1' WHERE folio_nota={folio_recuperar}")
                                            conn.commit()
                                            print('SE RECUPERÓ LA NOTA')
                                            break
                                elif opcion_recuperar=='NO':
                                    print('NO SE RECUPERÓ LA NOTA SELECCIONADA')
                                    break
                                else:
                                    print('Opción inválida. Escriba "SI" o "NO". Inténtelo de nuevo.')
                                    continue
                except Exception as e:
                    print('Ocurrió un error: {e} ')
                    continue

        def consulta_por_periodo_notas():
            while True:
                try:
                    with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                        mi_cursor= conn.cursor()

                        fecha_inicial = input("Ingrese la fecha inicial (MM-DD-AAAA) o presione Enter para usar la fecha 01-01-2000: ")
                        if fecha_inicial.strip() == '':
                            fecha_inicial = datetime.date(2000, 1, 1)
                        else:
                            fecha_inicial = datetime.datetime.strptime(fecha_inicial, "%m-%d-%Y").date()

                        fecha_final = input("Ingrese la fecha final (MM-DD-AAAA) o presione Enter para usar la fecha actual: ")
                        if fecha_final.strip() == '':
                            fecha_final = datetime.date.today()
                        else:
                            fecha_final = datetime.datetime.strptime(fecha_final, "%m-%d-%Y").date()

                        if fecha_final < fecha_inicial:
                            print(f"\nLA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL, INTENTE DE NUEVO.")
                            continue

                        mi_cursor.execute(f"SELECT * FROM notas WHERE fecha_nota BETWEEN '{fecha_inicial}' AND '{fecha_final}' ")
                        notas_periodo=mi_cursor.fetchall()

                        if not notas_periodo:
                                print('No hay notas creadas en ese periodo.')
                                break
                        else:
                            print(f"\nNotas encontradas en el período: ")
                            print(f"\nFolio nota\tFecha nota\tClave cliente\tTotal nota\tEstatus")
                            for nota in notas_periodo:
                                print(f"{nota[0]}\t\t{nota[1]}\t\t{nota[2]}\t{nota[3]}\t\t{nota[4]}")

                            mi_cursor.execute(f"SELECT total_nota FROM notas WHERE fecha_nota BETWEEN '{fecha_inicial}' AND '{fecha_final}' ")
                            notas_periodo_monto_promedio=mi_cursor.fetchall()

                            monto_notas=0
                            for costo in notas_periodo_monto_promedio:
                                monto_notas= monto_notas + int(costo[0])

                            promedio_notas=monto_notas / len(notas_periodo_monto_promedio)
                            print(f"\nEl monto promedio de las notas en el período es de: ${promedio_notas:.2f}")

                            print(f"\nOpciones de exportación:")
                            print(f"1. Excel")
                            print(f"2. CSV")
                            print(f"3. Volver a menú de reportes")



                            opcion_exportar=int(input(f'\nOpción a elegir: '))
                            match opcion_exportar:
                                case 1:
                                    exportar_a_excel_notas(fecha_inicial, fecha_final, promedio_notas)
                                    break

                                case 2:
                                    fecha_inicial_db = fecha_inicial.strftime("%m-%d-%Y")
                                    fecha_final_db = fecha_final.strftime("%m-%d-%Y")

                                    nombre_archivo_csv=f"ReportePorPeriodo_{fecha_inicial_db}_{fecha_final_db}.csv"
                                    with open(nombre_archivo_csv, 'w', newline='') as file:
                                        writer = csv.writer(file)
                                        writer.writerow(["Folio nota", "Fecha nota", "Clave Cliente", "Total nota", "Estatus"])

                                        for nota in notas_periodo:
                                            writer.writerow([nota[0], nota[1], nota[2], nota[3], nota[4]])

                                        writer.writerow(["Promedio de las notas en el periodo: ", promedio_notas])

                                        print(f"Se ha exportado la información a '{nombre_archivo_csv}'.")
                                        break

                                case 3:
                                    print('Volviendo al menú de reportes.')
                                    break
                                case _:
                                    print("Opción no válida. Por favor, ingrese un número del 1 al 3.")

                except ValueError:
                    print("Formato de fecha incorrecto.")
                    continue

                except Exception as e:
                    print(f"Ocurrió un error: {str(e)}")



        def exportar_a_excel_notas(fecha_inicial, fecha_final, promedio_notas):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Notas_Periodo"


            sheet['A1'] = "Folio nota"
            sheet['B1'] = "Fecha nota"
            sheet['C1'] = "Clave Cliente"
            sheet['D1'] = "Total nota"
            sheet['E1'] = "Estatus"
            sheet['F1'] = "Promedio notas en el período"

            row = 2
            with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                mi_cursor= conn.cursor()
                mi_cursor.execute(f"SELECT * FROM notas WHERE fecha_nota BETWEEN '{fecha_inicial}' AND '{fecha_final}' ")
                notas_periodo_db=mi_cursor.fetchall()
                print (notas_periodo_db)
                for nota in notas_periodo_db:
                    sheet[f'A{row}'] = nota[0]
                    sheet[f'B{row}'] = nota[1]
                    sheet[f'C{row}'] = nota[2]
                    sheet[f'D{row}'] = nota[3]
                    sheet[f'E{row}'] = nota[4]
                    row += 1
                sheet[f'F{row}'] = promedio_notas

            fecha_inicial_db = fecha_inicial.strftime("%m-%d-%Y")
            fecha_final_db = fecha_final.strftime("%m-%d-%Y")

            nombre_archivo_excel = f"ReportePorPeriodo_{fecha_inicial_db}_{fecha_final_db}.xlsx"
            workbook.save(nombre_archivo_excel)
            print(f"Se ha exportado la información a '{nombre_archivo_excel}'.")

        def consultar_por_folio_notas():
            with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                mi_cursor= conn.cursor()
                mi_cursor.execute(f"SELECT notas.folio_nota, notas.fecha_nota, clientes.nombre_cliente FROM notas INNER JOIN clientes ON notas.clave_cliente=clientes.clave_cliente ORDER BY notas.folio_nota ")
                notas_folio=mi_cursor.fetchall()
                print(f"\nCONSULTA POR FOLIO: ")
                print(f"Folio nota\tFecha nota\tNombre cliente")
                for nota in notas_folio:
                    print(f"{nota[0]}\t\t{nota[1]}\t{nota[2]}")
                while True:
                    try:
                        folio_consulta_notas=int(input(f'\nIngrese el folio de la nota a consultar / Si no se va a consultar una nota escriba 0: '))
                        if folio_consulta_notas == 0:
                            print('Usted ha decidido no consultar ninguna nota')
                            break
                        if not folio_consulta_notas:
                            print('El folio no se puede quedar en blanco. Por favor, ingrese un folio válido.')
                            continue

                        with sqlite3.connect ('Evidencia3_Prueba.db') as conn:
                            mi_cursor= conn.cursor()
                            mi_cursor.execute(f"SELECT notas.folio_nota, notas.fecha_nota, clientes.clave_cliente, clientes.nombre_cliente, clientes.RFC, clientes.correo_cliente, detalle_notas.id_detalle, servicios.nombre_servicio, servicios.costo_servicio, notas.estatus \
                                                FROM notas \
                                                INNER JOIN clientes ON notas.clave_cliente = clientes.clave_cliente\
                                                INNER JOIN detalle_notas ON notas.folio_nota = detalle_notas.folio_nota\
                                                INNER JOIN servicios ON detalle_notas.clave_servicio = servicios.clave_servicio WHERE notas.folio_nota={folio_consulta_notas} AND notas.estatus='1' ")

                            notas_folio=mi_cursor.fetchall()
                            if folio_consulta_notas not in [nota[0] for nota in notas_folio]:
                                print('La nota no se encuentra en el sistema. Por favor, ingrese un folio válido.')
                                continue
                            else:
                                print("\n{:<12}\t{:<14}\t{:<13}\t{:<16}\t{:<11}\t{:<19}\t{:<15}\t\t{:<20}\t{:<15}\t{:<12}".format("Folio nota", "Fecha nota", "Clave cliente", "Nombre cliente", "RFC cliente", "Correo cliente", "ID detalle nota", "Servicio realizado", "Costo servicio", "Estatus nota"))
                                for nota in notas_folio:
                                    print("{:<12}\t{:<14}\t{:<13}\t{:<16}\t{:<11}\t{:<19}\t{:<15}\t\t{:<20}\t${:.2f}\t\t{:<12}".format(nota[0], nota[1], nota[2], nota[3], nota[4], nota[5], nota[6], nota[7], nota[8], nota[9]))
                                break

                    except ValueError:
                        print('Ingrese un numero valido. Intenta de nuevo.')


 
